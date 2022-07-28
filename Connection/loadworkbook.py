import pandas as pd
import openpyxl
import glob
from tqdm import tqdm
from App import Transfer
import os
import yaml

# ============= #
# === INPUT === #
# ============= #

tf = Transfer()
conn = tf.database_connection()

#open pre-defined configurations
with open("config.yaml", "r") as stream:
    config = yaml.safe_load(stream)

#source inputs from yaml file, replace with path to data file
csv_input = config['folders']['csv_filename']


# ----------------------------------------------------------------------------------------------

# ================= #
# === FUNCTIONS === #
# ================= #

def load_workbook():
    for f in glob.glob(tf.working_filepath(csv_input)):
    #Load excel file
        wb = openpyxl.load_workbook(filename=f, read_only=True)
        #Load each sheet in workbook
        for sheet in tqdm(wb.sheetnames):
            ws = wb[sheet]
        #Load rows
            rows = ws.rows
            first_row = [cell.value for cell in next(rows)]
        #Load data
            data = []
            for row in rows:
                record = {}
                for key, cell in zip(first_row, row):
                    record[key] = cell.value
                data.append(record)
            df = pd.DataFrame(data)
            #Load sql
            df.to_sql(f'{sheet}', conn, schema='input', index=False, if_exists='replace')


load_workbook()